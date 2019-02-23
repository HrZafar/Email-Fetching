import email
import imaplib
from datetime import *
import calendar
from openpyxl import *
import os.path


def get_date(item):
    # returns date of email in string format
    res, data = mail.uid('fetch', item, '(RFC822)')
    raw = data[0][1].decode("utf-8")
    e_msg = email.message_from_string(raw)
    return e_msg['Date'][5:16]


def get_strDate(date_date):
    # converts datetime object in desired string format
    date_date = str(date_date)
    date_date = date_date[0:11]
    date_date = date_date.split('-')
    date_date = date_date[-1].strip() + '-' + calendar.month_abbr[int(date_date[1])] + '-' + date_date[0]
    return date_date


# creating excel file if it doesn't exist
if not (os.path.exists('email_ids.xlsx')):
    wb = Workbook()
    wb.create_sheet("email_ids.xlsx")
    wb.save('email_ids.xlsx')

# loading excel file
wb = load_workbook('email_ids.xlsx')
sheet = wb.active
sheet.column_dimensions['A'].width = 30
sheet.column_dimensions['B'].width = 30
sheet.column_dimensions['C'].width = 30
sheet.column_dimensions['D'].width = 60
sheet.cell(row=1, column=1).value = 'Date'
sheet.cell(row=1, column=2).value = 'Sender Name'
sheet.cell(row=1, column=3).value = 'Sender Email Id'
sheet.cell(row=1, column=4).value = 'To'
wb.save('email_ids.xlsx')

username = "example@gmail.com"
password = "password"

mail = imaplib.IMAP4_SSL("imap.gmail.com")
mail.login(username, password)
mail.select('inbox')

# getting all email from inbox
result, data = mail.uid('search', None, 'ALL')
inbox_items = data[0].split()

# getting oldest and newest email dates
oldest_date = get_date(inbox_items[0])
newest_date = get_date(inbox_items[-1])

# converting string date into datetime object
oldest_date = datetime.strptime(oldest_date, '%d %b %Y')
newest_date = datetime.strptime(newest_date, '%d %b %Y')

curr_date = newest_date
while curr_date >= oldest_date:
    # iterating through all dates between oldest and newest dates
    stringDate = get_strDate(curr_date)  # getting date in desired string format i.e. 01-Jan-2019
    result, data = mail.uid('search', None, 'SENTON ' + stringDate)
    inbox_items = data[0].split()

    # loading excel file
    wb = load_workbook('email_ids.xlsx')
    sheet = wb.active

    for i in range(len(inbox_items)):
        # iterating through all emails of current date
        result2, email_data = mail.uid('fetch', inbox_items[i], '(RFC822)')
        raw_email = email_data[0][1].decode("utf-8")
        msg = email.message_from_string(raw_email)

        # adding data to excel file
        current_row = sheet.max_row
        sheet.cell(row=current_row + 1, column=1).value = msg['Date']
        if '<' in msg['From']:
            sender=msg['From'].split('<')
            sender[1]=sender[1][0:-1]
            sheet.cell(row=current_row + 1, column=2).value = sender[0]
            sheet.cell(row=current_row + 1, column=3).value = sender[1]
        else:
            sheet.cell(row=current_row + 1, column=3).value = msg['From']
        sheet.cell(row=current_row + 1, column=4).value = msg['To']

    # saving excel file
    wb.save('email_ids.xlsx')

    curr_date = curr_date - timedelta(days=1)
