from googleapiclient.discovery import build
from httplib2 import Http
from oauth2client import file, client, tools
from openpyxl import *
import os.path
from datetime import *

SCOPES = 'https://www.googleapis.com/auth/gmail.readonly'


def get_value(head, val):
    for i in head:
        if i['name'] == val:
            return i['value']


# creating excel file if it doesn't exist
if not (os.path.exists('email_ids.xlsx')):
    wb = Workbook()
    wb.create_sheet("email_ids.xlsx")
    wb.save('email_ids.xlsx')

# loading excel file
wb = load_workbook('email_ids.xlsx')
sheet = wb.active
sheet.column_dimensions['A'].width = 30
sheet.column_dimensions['B'].width = 30
sheet.column_dimensions['C'].width = 30
sheet.column_dimensions['D'].width = 60
sheet.cell(row=1, column=1).value = 'Date'
sheet.cell(row=1, column=2).value = 'Sender Name'
sheet.cell(row=1, column=3).value = 'Sender Email Id'
sheet.cell(row=1, column=4).value = 'To'
wb.save('email_ids.xlsx')


def main():
    #getting rights to read gmail
    store = file.Storage('token.json')
    creds = store.get()
    if not creds or creds.invalid:
        flow = client.flow_from_clientsecrets('credentials.json', SCOPES)
        creds = tools.run_flow(flow, store)
    service = build('gmail', 'v1', http=creds.authorize(Http()))

    # getting all emails from inbox
    results = service.users().messages().list(userId='me', labelIds=['INBOX']).execute()
    messages = results.get('messages', [])

    # getting oldest and newest email dates
    newest_date = get_value(
        service.users().messages().get(userId='me', id=messages[0]['id']).execute()['payload']['headers'], 'Date')[5:16]
    oldest_date = get_value(
        service.users().messages().get(userId='me', id=messages[-1]['id']).execute()['payload']['headers'], 'Date')[
                  5:16]

    # converting string date into datetime object
    oldest_date = datetime.strptime(oldest_date, '%d %b %Y')
    newest_date = datetime.strptime(newest_date, '%d %b %Y')

    date_after = newest_date
    date_before = date_after + timedelta(days=1)
    while date_after >= oldest_date:
        # iterating through all dates between oldest and newest dates
        query = "after: {0} before: {1}".format(date_after.strftime('%Y/%m/%d'), date_before.strftime('%Y/%m/%d'))
        results = service.users().messages().list(userId='me', labelIds=['INBOX'], q=query).execute()
        messages = results.get('messages', [])

        if messages:
            # loading excel file
            wb = load_workbook('email_ids.xlsx')
            sheet = wb.active

            for message in messages:
                # iterating through all emails of current date
                msg = service.users().messages().get(userId='me', id=message['id']).execute()
                header = msg['payload']['headers']

                # adding data to excel file
                current_row = sheet.max_row
                sheet.cell(row=current_row + 1, column=1).value = get_value(header, 'Date')
                sender = get_value(header, 'From')
                if '<' in sender:
                    sender = sender.split('<')
                    sender[1] = sender[1][0:-1]
                    sheet.cell(row=current_row + 1, column=2).value = sender[0]
                    sheet.cell(row=current_row + 1, column=3).value = sender[1]
                else:
                    sheet.cell(row=current_row + 1, column=3).value = sender
                sheet.cell(row=current_row + 1, column=4).value = get_value(header, 'To')

            # saving excel file
            wb.save('email_ids.xlsx')

        date_after = date_after - timedelta(days=1)
        date_before = date_after + timedelta(days=1)


if __name__ == '__main__':
    main()
